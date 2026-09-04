import openpyxl
import warnings
import json

warnings.filterwarnings("ignore")

f = "/root/.claude/uploads/726a18de-9a86-57ac-9d50-9c356d0f76b4/5b573ae3-__________________________________.xlsx"
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb["СП"]

# Трёхслойные СП с наполнителем из минеральной ваты (105 кг/м3) — "применяем в стандартном расчёте".
data = []
for r in range(9, 16):
    thickness = ws.cell(row=r, column=2).value
    if thickness is None:
        continue
    data.append(
        {
            "thickness_mm": thickness,
            "wallPriceZLock": ws.cell(row=r, column=3).value,
            "wallPriceSecretFix": ws.cell(row=r, column=4).value,
            "roofPrice": ws.cell(row=r, column=5).value,
            "weight_kg_m2": ws.cell(row=r, column=6).value,
        }
    )

with open("/home/user/SprintM/data/sandwich_panel_prices.json", "w", encoding="utf-8") as fh:
    json.dump(data, fh, ensure_ascii=False, indent=2)
print(json.dumps(data, ensure_ascii=False, indent=2))
