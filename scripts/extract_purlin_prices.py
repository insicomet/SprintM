import openpyxl
import warnings
import json

warnings.filterwarnings("ignore")

f = "/root/.claude/uploads/726a18de-9a86-57ac-9d50-9c356d0f76b4/5b573ae3-__________________________________.xlsx"
wb = openpyxl.load_workbook(f, data_only=True)


def extract(sn, maxrow):
    ws = wb[sn]
    data = []
    for r in range(2, maxrow + 1):
        name = ws.cell(row=r, column=1).value
        if not isinstance(name, str) or not name.strip():
            continue
        data.append(
            {
                "name": name,
                "code": ws.cell(row=r, column=2).value,
                "group": ws.cell(row=r, column=3).value,
                "price0": ws.cell(row=r, column=4).value,
                "unit": ws.cell(row=r, column=5).value,
                "weight_kg": ws.cell(row=r, column=6).value,
                "pricePerTon": ws.cell(row=r, column=7).value,
                "priceSale": ws.cell(row=r, column=8).value,
            }
        )
    return data


sheets = {
    "ПС": 138,
    "ТПС": 341,
    "ПZ и TПZ": 80,
}
out = {}
for sn, maxrow in sheets.items():
    out[sn] = extract(sn, maxrow)
    print(sn, len(out[sn]))

with open("/home/user/SprintM/data/purlin_family_prices.json", "w", encoding="utf-8") as fh:
    json.dump(out, fh, ensure_ascii=False, indent=2)
print("DONE")
